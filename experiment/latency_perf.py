"""
Latency Performance Benchmarking Script

Compares inference latency across:
  - Qwen 0.6B / 4B / 8B
  × LoRA / DoRA / Shadow

A single MMLU record is used as the probe input.

CUDA_VISIBLE_DEVICES=6 python latency_perf.py \
  --ckpt_0.6B_lora   /home/lxm/workspace/Shadow/outputs/mmlu_suite_lora/mmlu_mmlu_full/lora/ \
  --ckpt_0.6B_dora   /home/lxm/workspace/Shadow/outputs/mmlu_suite_dora/mmlu_mmlu_full/dora/ \
  --ckpt_0.6B_shadow erin99/Qwen3-0.6B-MMLU-Shadow \
  --ckpt_4B_lora     /home/lxm/workspace/Shadow/outputs/mmlu_suite_lora_4B/mmlu_mmlu_full/lora/ \
  --ckpt_4B_dora     /home/lxm/workspace/Shadow/outputs/mmlu_suite_dora_4B/mmlu_mmlu_full/dora/ \
  --ckpt_4B_shadow   erin99/Qwen3-4B-MMLU-Shadow \
  --ckpt_8B_lora     /home/lxm/workspace/Shadow/outputs/mmlu_suite_lora_8B/mmlu_mmlu_full/lora/ \
  --ckpt_8B_dora     /home/lxm/workspace/Shadow/outputs/mmlu_suite_dora_8B/mmlu_mmlu_full/dora/ \
  --ckpt_8B_shadow   erin99/Qwen3-8B-MMLU-Shadow \
  --warmup_runs 3 --timed_runs 10 \
  --max_new_tokens 8 \
  --output_json  latency.json \
  --output_plot  latency.png \
  --output_excel latency.xlsx
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import torch
from datasets import load_dataset
from peft import PeftModel
from transformers import AutoModelForCausalLM, AutoTokenizer

# ---------------------------------------------------------------------------
# ShadowPEFT path setup (mirrors run_shadow_peft_ood_eval.py)
# ---------------------------------------------------------------------------
SHADOW_PEFT_PATH = Path(__file__).parent.parent / "ShadowPEFT" / "src"
if str(SHADOW_PEFT_PATH) not in sys.path:
    sys.path.insert(0, str(SHADOW_PEFT_PATH))

from shadow_peft import ShadowForCausalLM  # noqa: E402  (after path setup)

# ---------------------------------------------------------------------------
# Add Shadow directory to path so data_utils is importable
# ---------------------------------------------------------------------------
_SHADOW_DIR = Path(__file__).parent
if str(_SHADOW_DIR) not in sys.path:
    sys.path.insert(0, str(_SHADOW_DIR))

from data_utils import _build_chat_texts  # noqa: E402


# ---------------------------------------------------------------------------
# Canonical model names
# ---------------------------------------------------------------------------
MODEL_NAMES = {
    "0.6B": "Qwen/Qwen3-0.6B",
    "4B":   "Qwen/Qwen3-4B",
    "8B":   "Qwen/Qwen3-8B",
}

MODES = ("lora", "dora", "shadow")

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Latency benchmark: LoRA vs DoRA vs Shadow across Qwen model sizes."
    )

    # ── checkpoint directories ─────────────────────────────────────────────
    for size in MODEL_NAMES:
        for mode in MODES:
            p.add_argument(
                f"--ckpt_{size}_{mode}",
                type=str,
                default=None,
                metavar="DIR",
                help=f"Checkpoint dir for Qwen-{size} {mode} (skip if None).",
            )

    # ── model / hardware ───────────────────────────────────────────────────
    p.add_argument("--attn_implementation", default="flash_attention_2",
                   choices=("sdpa", "eager", "flash_attention_2"))
    p.add_argument("--bf16", type=int, default=1, choices=[0, 1])
    p.add_argument("--fp16", type=int, default=0, choices=[0, 1])
    p.add_argument("--device", type=str, default="cuda",
                   help="Torch device (e.g. 'cuda', 'cuda:0', 'cpu').")

    # ── MMLU probe ─────────────────────────────────────────────────────────
    p.add_argument("--mmlu_subset", type=str, default="high_school_mathematics",
                   help="MMLU subset to draw the probe from.")
    p.add_argument("--mmlu_split", type=str, default="test",
                   choices=("test", "validation", "dev"))
    p.add_argument("--probe_index", type=int, default=0,
                   help="Index of the MMLU record to use as probe.")
    p.add_argument("--max_seq_length", type=int, default=512)

    # ── generation ─────────────────────────────────────────────────────────
    p.add_argument("--max_new_tokens", type=int, default=8,
                   help="Max new tokens to generate per probe call.")

    # ── timing ─────────────────────────────────────────────────────────────
    p.add_argument("--warmup_runs", type=int, default=3,
                   help="Number of warmup forward passes before timing.")
    p.add_argument("--timed_runs", type=int, default=10,
                   help="Number of timed forward passes to average.")

    # ── misc ───────────────────────────────────────────────────────────────
    p.add_argument("--seed", type=int, default=42)
    p.add_argument("--output_json", type=str, default=None,
                   help="Optional path to save results as JSON.")
    p.add_argument("--output_plot", type=str, default=None,
                   help="Optional path to save the latency chart (e.g. results/latency.png).")
    p.add_argument("--output_excel", type=str, default=None,
                   help="Optional path to save an Excel workbook (e.g. results/latency.xlsx).")

    # ShadowPEFT inference knobs
    p.add_argument("--inference_mode", default="base_shadow",
                   choices=("base_shadow", "shadow_only"))
    p.add_argument("--shadow_loss_weight", type=float, default=0.05)

    return p.parse_args()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def set_seed(seed: int) -> None:
    import random
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)


def _get_dtype(args: argparse.Namespace) -> torch.dtype:
    if args.bf16:
        return torch.bfloat16
    if args.fp16:
        return torch.float16
    return torch.float32


# ---------------------------------------------------------------------------
# MMLU probe builder (no HF Datasets processing, direct raw access)
# ---------------------------------------------------------------------------

def build_mmlu_probe(
    subset: str,
    split: str,
    probe_index: int,
    tokenizer: AutoTokenizer,
    max_length: int,
    device: torch.device,
) -> Tuple[torch.Tensor, torch.Tensor, str, str]:
    """
    Load one raw MMLU example and return:
        (input_ids, attention_mask, question_text, gold_answer_letter)

    The prompt format matches data_utils.load_mmlu_dataset exactly so that
    latency numbers are representative of real evaluation workloads.
    """
    raw_ds = load_dataset(
        "cais/mmlu", subset, split=split,
        download_mode="reuse_dataset_if_exists",
    )

    # Unwrap potential nested structure (some MMLU splits wrap rows differently)
    if "question" not in raw_ds.column_names and "test" in raw_ds.column_names:
        raw_ds = raw_ds["test"]

    example = raw_ds[probe_index]
    question   = example["question"]
    choices    = example["choices"]
    answer_idx = int(example["answer"])

    choice_letters = ["A", "B", "C", "D"]
    answer_letter  = choice_letters[answer_idx]

    # Build user content (same format as data_utils.load_mmlu_dataset)
    user_content  = f"Question: {question}\n\nOptions:\n"
    for i, ch in enumerate(choices):
        user_content += f"{choice_letters[i]}: {ch}\n"
    user_content += (
        "\nInstructions: Answer with ONLY the letter (A, B, C, or D). "
        "Do not include any explanation, reasoning, or additional text. "
        "Answer:\n\n"
    )

    # Build prompt text via chat template (enable_thinking=False like data_utils)
    _, prompt_text = _build_chat_texts(tokenizer, user_content, answer_letter)

    enc = tokenizer(
        prompt_text,
        truncation=True,
        max_length=max_length,
        padding=False,
        return_tensors="pt",
    )

    input_ids      = enc["input_ids"].to(device)
    attention_mask = enc["attention_mask"].to(device)

    return input_ids, attention_mask, question, answer_letter


# ---------------------------------------------------------------------------
# Model loaders
# ---------------------------------------------------------------------------

def _prepare_tokenizer(model_name: str) -> AutoTokenizer:
    tok = AutoTokenizer.from_pretrained(
        model_name, trust_remote_code=True, use_fast=True
    )
    if tok.pad_token is None:
        tok.pad_token = tok.eos_token
    return tok


def _prepare_base_model(
    model_name: str,
    attn_impl: str,
    dtype: torch.dtype,
) -> AutoModelForCausalLM:
    return AutoModelForCausalLM.from_pretrained(
        model_name,
        torch_dtype=dtype,
        attn_implementation=attn_impl,
        trust_remote_code=True,
    )


def load_lora_or_dora_model(
    model_name: str,
    checkpoint_dir: str,
    attn_impl: str,
    dtype: torch.dtype,
    tokenizer: AutoTokenizer,
    device: torch.device,
) -> PeftModel:
    base = _prepare_base_model(model_name, attn_impl, dtype)
    if tokenizer.pad_token_id != base.config.pad_token_id:
        base.config.pad_token_id = tokenizer.pad_token_id
    model = PeftModel.from_pretrained(base, checkpoint_dir)
    model = model.to(dtype).to(device)
    model.eval()
    return model


def load_shadow_model(
    model_name: str,
    checkpoint_dir: str,
    attn_impl: str,
    dtype: torch.dtype,
    tokenizer: AutoTokenizer,
    device: torch.device,
    inference_mode: str,
    shadow_loss_weight: float,
) -> ShadowForCausalLM:
    base = _prepare_base_model(model_name, attn_impl, dtype)
    if tokenizer.pad_token_id != base.config.pad_token_id:
        base.config.pad_token_id = tokenizer.pad_token_id
    model = ShadowForCausalLM.from_pretrained(
        base,
        checkpoint_dir,
        is_trainable=False,
        shadow_loss_weight=shadow_loss_weight,
        inference_mode=inference_mode,
    )
    model = model.to(dtype).to(device)
    model.eval()
    return model


# ---------------------------------------------------------------------------
# Timing helpers
# ---------------------------------------------------------------------------

def _sync(device: torch.device) -> None:
    """Synchronise CUDA / CPU before timing."""
    if device.type == "cuda":
        torch.cuda.synchronize(device)


def measure_latency(
    model,
    input_ids: torch.Tensor,
    attention_mask: torch.Tensor,
    max_new_tokens: int,
    warmup_runs: int,
    timed_runs: int,
    device: torch.device,
    pad_token_id: int,
    eos_token_id: int,
) -> Dict[str, float]:
    """
    Run *warmup_runs* warmup passes then *timed_runs* timed generate() calls.

    Returns dict with keys: mean_ms, std_ms, min_ms, max_ms, median_ms.
    """
    gen_kwargs = dict(
        input_ids=input_ids,
        attention_mask=attention_mask,
        max_new_tokens=max_new_tokens,
        do_sample=False,
        pad_token_id=pad_token_id,
        eos_token_id=eos_token_id,
    )

    with torch.inference_mode():
        # Warmup
        for _ in range(warmup_runs):
            _sync(device)
            model.generate(**gen_kwargs)
            _sync(device)

        # Timed runs
        times_ms: List[float] = []
        for _ in range(timed_runs):
            _sync(device)
            t0 = time.perf_counter()
            model.generate(**gen_kwargs)
            _sync(device)
            t1 = time.perf_counter()
            times_ms.append((t1 - t0) * 1_000.0)

    arr = np.array(times_ms)
    return {
        "mean_ms":   float(arr.mean()),
        "std_ms":    float(arr.std()),
        "min_ms":    float(arr.min()),
        "max_ms":    float(arr.max()),
        "median_ms": float(np.median(arr)),
        "runs":      timed_runs,
        "raw_ms":    times_ms,
    }


# ---------------------------------------------------------------------------
# Pretty-print helpers
# ---------------------------------------------------------------------------

_COL_WIDTH = 14
_LABEL_WIDTH = 30


def _fmt(val: Optional[float]) -> str:
    if val is None:
        return "N/A".center(_COL_WIDTH)
    return f"{val:>10.2f} ms"


def print_table(results: Dict[str, Dict]) -> None:
    """Print a formatted comparison table to stdout."""
    # Build row keys ordered by (size, mode)
    size_order = {"0.6B": 0, "4B": 1, "8B": 2}
    mode_order = {"lora": 0, "dora": 1, "shadow": 2}

    rows = sorted(
        results.keys(),
        key=lambda k: (size_order.get(k.split("_")[0], 99),
                       mode_order.get(k.split("_")[1], 99)),
    )

    header = (
        f"{'Model':<{_LABEL_WIDTH}}"
        f"{'Mean':>{_COL_WIDTH}}"
        f"{'Median':>{_COL_WIDTH}}"
        f"{'Std':>{_COL_WIDTH}}"
        f"{'Min':>{_COL_WIDTH}}"
        f"{'Max':>{_COL_WIDTH}}"
    )
    sep = "-" * len(header)

    print("\n" + "=" * len(header))
    print("LATENCY BENCHMARK RESULTS (MMLU probe, generate())")
    print("=" * len(header))
    print(header)
    print(sep)

    prev_size = None
    for key in rows:
        size, mode = key.split("_", 1)
        if prev_size and size != prev_size:
            print(sep)
        prev_size = size

        label = f"Qwen-{size} {mode.upper()}"
        stats = results[key]
        if stats is None:
            print(f"{label:<{_LABEL_WIDTH}}{'[SKIPPED]':>{_COL_WIDTH * 5}}")
            continue

        if "error" in stats:
            err = stats["error"][:40]
            print(f"{label:<{_LABEL_WIDTH}}{'[ERROR] ' + err:>{_COL_WIDTH * 5}}")
            continue

        print(
            f"{label:<{_LABEL_WIDTH}}"
            f"{_fmt(stats['mean_ms'])}"
            f"{_fmt(stats['median_ms'])}"
            f"{_fmt(stats['std_ms'])}"
            f"{_fmt(stats['min_ms'])}"
            f"{_fmt(stats['max_ms'])}"
        )

    print("=" * len(header))
    print(f"Warmup runs: {next(v['runs'] for v in results.values() if v and 'runs' in v)}")
    print()


# ---------------------------------------------------------------------------
# Fancy latency chart
# ---------------------------------------------------------------------------

# Palette: LoRA=indigo, DoRA=teal, Shadow=coral
_MODE_COLORS  = {"lora": "#6C8EBF", "dora": "#67AB9F", "shadow": "#647687"}
_MODE_LABELS  = {"lora": "LoRA", "dora": "DoRA", "shadow": "Shadow"}
_SIZE_ORDER   = ["0.6B", "4B", "8B"]
_MODE_ORDER   = ["lora", "dora", "shadow"]


def plot_results(results: Dict[str, object], output_path: str) -> None:
    """
    Save two separate white-background figures derived from *output_path*:

      <dir>/latency_inference.png  — Grouped bar chart (mean ± std)
      <dir>/latency_per_run.png    — Violin + jitter strip per model/mode
    """
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.patches as mpatches
    except ImportError:
        print("[plot] matplotlib not installed — skipping chart.")
        return

    out_dir = os.path.dirname(os.path.abspath(output_path))
    os.makedirs(out_dir, exist_ok=True)
    path_inference = os.path.join(out_dir, "latency_inference.png")
    path_per_run   = os.path.join(out_dir, "latency_per_run.png")

    # ── collect data ───────────────────────────────────────────────────────
    sizes_present = [s for s in _SIZE_ORDER
                     if any(f"{s}_{m}" in results and
                            isinstance(results[f"{s}_{m}"], dict) and
                            "mean_ms" in results[f"{s}_{m}"]
                            for m in _MODE_ORDER)]
    if not sizes_present:
        print("[plot] No valid results to plot.")
        return

    # shared helpers
    n_modes   = len(_MODE_ORDER)
    bar_w     = 0.22
    group_gap = bar_w * n_modes + 0.15
    x_centers = np.arange(len(sizes_present)) * (group_gap + 0.1)

    def _white_ax(ax):
        ax.set_facecolor("white")
        for spine in ax.spines.values():
            spine.set_edgecolor("#CCCCCC")

    # ── Figure 1: Inference Latency bar chart ─────────────────────────────
    fig1, ax_bar = plt.subplots(
        figsize=(max(8, len(sizes_present) * 3.2), 5),
        facecolor="white",
    )
    _white_ax(ax_bar)

    # Collect strip data while building bars
    strip_labels: List[str] = []
    strip_values: List[float] = []
    strip_colors: List[str]  = []

    for mi, mode in enumerate(_MODE_ORDER):
        offset = (mi - (n_modes - 1) / 2) * (bar_w + 0.03)
        means, stds, xs = [], [], []

        for si, size in enumerate(sizes_present):
            key  = f"{size}_{mode}"
            stat = results.get(key)
            if not isinstance(stat, dict) or "mean_ms" not in stat:
                continue
            xs.append(x_centers[si] + offset)
            means.append(stat["mean_ms"])
            stds.append(stat["std_ms"])
            for v in stat.get("raw_ms", [stat["mean_ms"]]):
                strip_labels.append(f"Qwen-{size}\n{_MODE_LABELS[mode]}")
                strip_values.append(v)
                strip_colors.append(_MODE_COLORS[mode])

        if not xs:
            continue

        ax_bar.bar(
            xs, means,
            width=bar_w,
            color=_MODE_COLORS[mode],
            alpha=0.85,
            label=_MODE_LABELS[mode],
            zorder=3,
            edgecolor="white",
            linewidth=0.8,
        )
        ax_bar.errorbar(
            xs, means, yerr=stds,
            fmt="none",
            ecolor="#555555",
            elinewidth=1.2,
            capsize=4,
            capthick=1.2,
            zorder=4,
        )
        # Value labels above each bar (placed after ylim is known via autoscale)
        for x, mean, std in zip(xs, means, stds):
            ax_bar.text(
                x, mean + std,
                f"{mean:.1f}",
                ha="center", va="bottom",
                fontsize=7.5, color="#222222",
                fontweight="bold",
                zorder=5,
            )

    ax_bar.set_xticks(x_centers)
    ax_bar.set_xticklabels([f"Qwen-{s}" for s in sizes_present],
                           color="#222222", fontsize=11, fontweight="bold")
    ax_bar.set_ylabel("Latency (ms)", color="#222222", fontsize=11)
    ax_bar.tick_params(colors="#444444")
    ax_bar.yaxis.grid(True, color="#E5E5E5", linewidth=0.7, zorder=0)
    ax_bar.set_axisbelow(True)
    # extend y-axis top by 15 % to give value labels breathing room
    ylo, yhi = ax_bar.get_ylim()
    ax_bar.set_ylim(ylo, yhi * 1.15)

    ax_bar.legend(
        handles=[mpatches.Patch(color=_MODE_COLORS[m], label=_MODE_LABELS[m])
                 for m in _MODE_ORDER],
        loc="upper left",
        framealpha=0.7,
        facecolor="white",
        edgecolor="#CCCCCC",
        fontsize=10,
    )

    fig1.tight_layout()
    fig1.savefig(path_inference, dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig1)
    print(f"[plot] Inference latency chart → {path_inference}")

    # ── Figure 2: Per-run distribution (violin + jitter) ──────────────────
    if not strip_labels:
        print("[plot] No raw run data for per-run chart — skipping.")
        return

    unique_labels: List[str] = []
    seen: set = set()
    for lbl in strip_labels:
        if lbl not in seen:
            unique_labels.append(lbl)
            seen.add(lbl)

    label_to_idx = {lbl: i for i, lbl in enumerate(unique_labels)}
    xs_strip     = [label_to_idx[lbl] for lbl in strip_labels]

    grouped: Dict[int, List[float]] = {}
    grouped_colors: Dict[int, str]  = {}
    for idx, val, col in zip(xs_strip, strip_values, strip_colors):
        grouped.setdefault(idx, []).append(val)
        grouped_colors[idx] = col

    vp_positions = sorted(grouped.keys())
    vp_data      = [grouped[i] for i in vp_positions]

    fig2, ax_viol = plt.subplots(
        figsize=(max(8, len(unique_labels) * 1.1), 5),
        facecolor="white",
    )
    _white_ax(ax_viol)

    if len(vp_data[0]) >= 3:
        parts = ax_viol.violinplot(
            vp_data, positions=vp_positions,
            widths=0.55, showmedians=True, showextrema=False,
        )
        for i, body in enumerate(parts["bodies"]):
            body.set_facecolor(grouped_colors[vp_positions[i]])
            body.set_alpha(0.30)
            body.set_edgecolor("none")
        parts["cmedians"].set_color("#333333")
        parts["cmedians"].set_linewidth(1.8)

    rng = np.random.default_rng(0)
    jitter = rng.uniform(-0.12, 0.12, size=len(xs_strip))
    ax_viol.scatter(
        np.array(xs_strip, dtype=float) + jitter,
        strip_values,
        c=strip_colors,
        s=24,
        alpha=0.75,
        zorder=4,
        edgecolors="none",
    )

    ax_viol.set_xticks(list(range(len(unique_labels))))
    ax_viol.set_xticklabels(unique_labels, color="#222222", fontsize=8.5)
    ax_viol.set_ylabel("Latency (ms)", color="#222222", fontsize=10)
    ax_viol.tick_params(colors="#444444")
    ax_viol.yaxis.grid(True, color="#E5E5E5", linewidth=0.7, zorder=0)
    ax_viol.set_axisbelow(True)

    fig2.tight_layout()
    fig2.savefig(path_per_run, dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig2)
    print(f"[plot] Per-run distribution chart → {path_per_run}")


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_excel(results: Dict[str, object], output_path: str) -> None:
    """
    Write a two-sheet Excel workbook:

    Sheet 1 "Summary"  — one row per (model_size, mode) with
                         mean / median / std / min / max (ms) plus status.
    Sheet 2 "Raw Runs" — one column per (model_size, mode) with every
                         individual timed run value, so users can build
                         their own pivot charts inside Excel.
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Alignment, Border, Font, PatternFill, Side,
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[excel] openpyxl not installed — skipping Excel export.\n"
              "        Install with:  pip install openpyxl")
        return

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    wb = openpyxl.Workbook()

    # ── colour palette ────────────────────────────────────────────────────
    _HDR_FILL   = PatternFill("solid", fgColor="1A1A2E")   # dark navy header
    _HDR_FONT   = Font(bold=True, color="FFFFFF", size=11)
    _SUBHDR_FILLS = {                                        # mode-tinted rows
        "lora":   PatternFill("solid", fgColor="D6E4F7"),
        "dora":   PatternFill("solid", fgColor="D6EEE9"),
        "shadow": PatternFill("solid", fgColor="FAE3D9"),
    }
    _BOLD = Font(bold=True)
    _CENTER = Alignment(horizontal="center", vertical="center")
    _THIN = Side(style="thin", color="AAAAAA")
    _BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

    def _style_header(cell, text):
        cell.value = text
        cell.font  = _HDR_FONT
        cell.fill  = _HDR_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER

    def _style_cell(cell, value, fill=None, bold=False, fmt=None):
        cell.value = value
        if fill:
            cell.fill = fill
        if bold:
            cell.font = _BOLD
        cell.alignment = _CENTER
        cell.border = _BORDER
        if fmt:
            cell.number_format = fmt

    # ── helpers ───────────────────────────────────────────────────────────
    size_order = {"0.6B": 0, "4B": 1, "8B": 2}
    mode_order = {"lora": 0, "dora": 1, "shadow": 2}
    sorted_keys = sorted(
        results.keys(),
        key=lambda k: (size_order.get(k.split("_")[0], 99),
                       mode_order.get(k.split("_")[1], 99)),
    )

    # ── Sheet 1: Summary ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.freeze_panes = "A2"

    headers = ["Model Size", "Mode", "Status",
               "Mean (ms)", "Median (ms)", "Std (ms)", "Min (ms)", "Max (ms)", "Runs"]
    for ci, h in enumerate(headers, 1):
        _style_header(ws1.cell(row=1, column=ci), h)

    for ri, key in enumerate(sorted_keys, 2):
        size, mode = key.split("_", 1)
        stat = results[key]
        fill = _SUBHDR_FILLS.get(mode)

        _style_cell(ws1.cell(row=ri, column=1), f"Qwen-{size}", fill=fill, bold=True)
        _style_cell(ws1.cell(row=ri, column=2), mode.upper(), fill=fill)

        if stat is None:
            _style_cell(ws1.cell(row=ri, column=3), "SKIPPED", fill=fill)
            for ci in range(4, 10):
                _style_cell(ws1.cell(row=ri, column=ci), "—", fill=fill)
        elif "error" in stat:
            _style_cell(ws1.cell(row=ri, column=3), "ERROR", fill=fill)
            _style_cell(ws1.cell(row=ri, column=4), stat["error"][:80], fill=fill)
            for ci in range(5, 10):
                _style_cell(ws1.cell(row=ri, column=ci), "—", fill=fill)
        else:
            _style_cell(ws1.cell(row=ri, column=3), "OK", fill=fill)
            for ci, field in enumerate(
                ["mean_ms", "median_ms", "std_ms", "min_ms", "max_ms"], 4
            ):
                _style_cell(ws1.cell(row=ri, column=ci),
                             round(stat[field], 3), fill=fill, fmt='0.000')
            _style_cell(ws1.cell(row=ri, column=9), stat.get("runs", ""), fill=fill)

    # Column widths
    col_widths = [14, 10, 10, 14, 14, 12, 12, 12, 8]
    for ci, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 2: Raw Runs ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Raw Runs")

    valid_keys = [k for k in sorted_keys
                  if isinstance(results[k], dict) and "raw_ms" in results[k]]

    # Header row: one column per model/mode combination
    for ci, key in enumerate(valid_keys, 1):
        size, mode = key.split("_", 1)
        _style_header(ws2.cell(row=1, column=ci), f"Qwen-{size}\n{mode.upper()}")
        ws2.row_dimensions[1].height = 28

    # Data rows: run values
    max_runs = max((len(results[k]["raw_ms"]) for k in valid_keys), default=0)
    for ri in range(max_runs):
        for ci, key in enumerate(valid_keys, 1):
            raw = results[key]["raw_ms"]
            if ri < len(raw):
                cell = ws2.cell(row=ri + 2, column=ci)
                cell.value = round(raw[ri], 4)
                cell.alignment = _CENTER
                cell.border = _BORDER
                cell.number_format = '0.0000'

    # "Run #" label column on the left
    ws2.insert_cols(1)
    _style_header(ws2.cell(row=1, column=1), "Run #")
    for ri in range(max_runs):
        cell = ws2.cell(row=ri + 2, column=1)
        cell.value = ri + 1
        cell.alignment = _CENTER
        cell.border = _BORDER
        cell.font = Font(bold=True)

    for ci in range(1, len(valid_keys) + 2):
        ws2.column_dimensions[get_column_letter(ci)].width = 16

    wb.save(output_path)
    print(f"[excel] Workbook saved to {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    args = parse_args()
    set_seed(args.seed)

    device = torch.device(args.device if torch.cuda.is_available() else "cpu")
    dtype  = _get_dtype(args)

    print("=" * 60)
    print("MMLU Latency Performance Benchmark")
    print("=" * 60)
    print(f"Device : {device}")
    print(f"DType  : {dtype}")
    print(f"MMLU   : {args.mmlu_subset} / {args.mmlu_split}[{args.probe_index}]")
    print(f"Warmup : {args.warmup_runs}  Timed: {args.timed_runs}")
    print(f"Max new tokens: {args.max_new_tokens}")
    print("=" * 60 + "\n")

    # Build the list of (size, mode, checkpoint_dir) to benchmark
    configs: List[Tuple[str, str, Optional[str]]] = []
    for size in MODEL_NAMES:
        for mode in MODES:
            attr = f"ckpt_{size}_{mode}"
            ckpt = getattr(args, attr, None)
            configs.append((size, mode, ckpt))

    results: Dict[str, object] = {}

    # We may need different tokenizers per model size (same family, so one
    # tokenizer per size is enough; reuse across modes for that size).
    tokenizer_cache: Dict[str, AutoTokenizer] = {}
    probe_cache: Dict[str, Tuple] = {}  # keyed by model_name

    for size, mode, ckpt in configs:
        key = f"{size}_{mode}"
        model_name = MODEL_NAMES[size]

        if ckpt is None:
            print(f"[SKIP] Qwen-{size} {mode.upper()} — no checkpoint provided.\n")
            results[key] = None
            continue

        # Treat as a local path only when it starts with '/' or '.';
        # otherwise assume it is a HuggingFace Hub repo ID and let the
        # PEFT / ShadowPEFT loaders resolve it remotely.
        _is_local = ckpt.startswith("/") or ckpt.startswith(".")
        if _is_local and not os.path.isdir(ckpt):
            print(f"[SKIP] Qwen-{size} {mode.upper()} — local path not found: {ckpt}\n")
            results[key] = None
            continue

        print(f"─── Qwen-{size} {mode.upper()} ────────────────────────────")
        print(f"    ckpt : {ckpt}")

        # ── tokenizer ──────────────────────────────────────────────────────
        if model_name not in tokenizer_cache:
            print(f"    Loading tokenizer for {model_name} …")
            tokenizer_cache[model_name] = _prepare_tokenizer(model_name)
        tokenizer = tokenizer_cache[model_name]

        # ── probe ──────────────────────────────────────────────────────────
        if model_name not in probe_cache:
            print(f"    Building MMLU probe …")
            input_ids, attn_mask, question, gold = build_mmlu_probe(
                subset=args.mmlu_subset,
                split=args.mmlu_split,
                probe_index=args.probe_index,
                tokenizer=tokenizer,
                max_length=args.max_seq_length,
                device=device,
            )
            probe_cache[model_name] = (input_ids, attn_mask, question, gold)
            print(f"    Probe Q : {question[:80]}…")
            print(f"    Gold    : {gold}")
            print(f"    Tokens  : {input_ids.shape[-1]}")
        input_ids, attn_mask, _, _ = probe_cache[model_name]

        # ── model ──────────────────────────────────────────────────────────
        try:
            print(f"    Loading model …")
            if mode in ("lora", "dora"):
                model = load_lora_or_dora_model(
                    model_name, ckpt, args.attn_implementation,
                    dtype, tokenizer, device,
                )
            else:  # shadow
                model = load_shadow_model(
                    model_name, ckpt, args.attn_implementation,
                    dtype, tokenizer, device,
                    inference_mode=args.inference_mode,
                    shadow_loss_weight=args.shadow_loss_weight,
                )
        except Exception as exc:
            print(f"    [ERROR] model load failed: {exc}\n")
            results[key] = {"error": str(exc)}
            continue

        # ── timing ─────────────────────────────────────────────────────────
        try:
            print(f"    Benchmarking ({args.warmup_runs} warmup + {args.timed_runs} timed) …")
            stats = measure_latency(
                model=model,
                input_ids=input_ids,
                attention_mask=attn_mask,
                max_new_tokens=args.max_new_tokens,
                warmup_runs=args.warmup_runs,
                timed_runs=args.timed_runs,
                device=device,
                pad_token_id=tokenizer.pad_token_id,
                eos_token_id=tokenizer.eos_token_id,
            )
            print(
                f"    mean={stats['mean_ms']:.2f} ms  "
                f"median={stats['median_ms']:.2f} ms  "
                f"std={stats['std_ms']:.2f} ms"
            )
            results[key] = stats
        except Exception as exc:
            print(f"    [ERROR] timing failed: {exc}\n")
            results[key] = {"error": str(exc)}

        # ── cleanup ────────────────────────────────────────────────────────
        del model
        if device.type == "cuda":
            torch.cuda.empty_cache()
        print()

    # ── summary table ───────────────────────────────────────────────────────
    print_table(results)

    # ── optional chart ──────────────────────────────────────────────────────
    if args.output_plot:
        plot_results(results, args.output_plot)

    # ── optional Excel export ────────────────────────────────────────────────
    if args.output_excel:
        export_excel(results, args.output_excel)

    # ── optional JSON export ─────────────────────────────────────────────────
    if args.output_json:
        # Remove raw_ms list for cleaner JSON (can be large)
        export = {}
        for k, v in results.items():
            if isinstance(v, dict) and "raw_ms" in v:
                v = {kk: vv for kk, vv in v.items() if kk != "raw_ms"}
            export[k] = v
        with open(args.output_json, "w", encoding="utf-8") as f:
            json.dump(export, f, indent=2)
        print(f"Results saved to {args.output_json}")


if __name__ == "__main__":
    main()
